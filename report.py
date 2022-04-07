import geopandas as gpd
import openpyxl


def get_report(cadasters_js, fields_js):
    def get_area(gdf):
        return gdf['geometry'].to_crs({'proj':'cea'}).area / 10 ** 4

    try:
        cads = gpd.GeoDataFrame.from_features(cadasters_js)
        fields = gpd.GeoDataFrame.from_features(fields_js)
    except TypeError:
        return -1

    cads = cads[['Parcel_KN', 'geometry']]
    fields = fields[['Name', 'geometry']]

    entry = gpd.overlay(cads, fields, how='intersection') # ENTRY
    unused = gpd.overlay(cads, fields, how='difference') #UNUSED
    capture = gpd.overlay(fields, cads, how='difference')

    cads['area'] = get_area(cads)
    fields['area'] = get_area(fields)
    entry['area'] = get_area(entry)
    unused['area'] = get_area(unused)
    capture['area'] = get_area(capture)


    cads_area = cads.groupby('Parcel_KN')['area'].sum()
    fields_area = fields.groupby('Name')['area'].sum()
    unused_area = unused.groupby('Parcel_KN')['area'].sum()
    entry_area = entry.groupby(['Parcel_KN', 'Name'])['area'].sum()
    capture_area = capture.groupby(['Name'])['area'].sum()

    def find_col(text):
        i = 1
        while i <= sheet.max_column:
            if sheet.cell(row=1, column=i).value == text:
                return i
            i +=1

    def find_row(text):
        i = 1
        while i <= sheet.max_row:
            if sheet.cell(row=i, column=1).value == text:
                return i
            i +=1

    def add_col(text):
        nonlocal cur_col
        cur_col += 1
        cell = sheet.cell(row=1, column=cur_col)
        cell.value = text

    def add_row(text):
        nonlocal cur_row
        cur_row += 1
        cell = sheet.cell(row=cur_row, column=1)
        cell.value = text

    def insert_value(value, row, col):
        cell = sheet.cell(row=row, column=col)
        cell.value = value

    def square_per_col(col):
        s = 0
        for i in range(2, sheet.max_row - 2):
            val = sheet.cell(row=i, column=col).value
            if val is not None:
                s += val

        return s

    def square_per_row(row):
        s = 0
        for i in range(2, sheet.max_column - 2):
            val = sheet.cell(row=row, column=i).value
            if val is not None:
                s += val
        return s

    wb = openpyxl.workbook.Workbook()
    sheet = wb.active


    sheet['A1'].value = 'Поля/Кад. н.'
    cur_row = sheet.max_row
    cur_col = sheet.max_column
    for cad_n in cads_area.index:
        add_col(cad_n)

    for field in fields_area.index:
        add_row(field)

    for stat in ['Площадь', 'Площадь пересечения']:
        add_col(stat)
        add_row(stat)
    add_row('Неиспольз.')
    add_col('Захват')

    cur_row = find_row('Неиспольз.')
    for cad_n in cads_area.index:
        cur_col = find_col(cad_n)
        insert_value(unused_area[cad_n], cur_row, cur_col)

    cur_col = find_col('Захват')
    for field_name in capture_area.index:
        cur_row = find_row(field_name)
        insert_value(capture_area[field_name], cur_row, cur_col)

    cur_row = find_row('Площадь')
    for cad_n in cads_area.index:
        cur_col = find_col(cad_n)
        value =cads_area[cad_n]
        insert_value(value, cur_row, cur_col)

    cur_col = find_col('Площадь')
    for field_name in fields_area.index:
        cur_row = find_row(field_name)
        insert_value(fields_area[field_name], cur_row, cur_col)

    for cad_n, field_name in entry_area.index:
        cur_col = find_col(cad_n)
        cur_row = find_row(field_name)
        insert_value(entry_area[cad_n][field_name], cur_row, cur_col)

    for i in range(2, sheet.max_column - 2):
        insert_value(square_per_col(i), col=i, row=sheet.max_row-1)
    for i in range(2, sheet.max_row - 2):
        insert_value(square_per_row(i), col=sheet.max_column -1, row=i)

    return wb

